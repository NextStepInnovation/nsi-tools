import re
import tempfile
from pathlib import Path, PurePosixPath
import logging

from webdav.client import Client

from .toolz import *

log = logging.getLogger(__name__)
log.addHandler(logging.NullHandler())

class DavPath:
    client_function = None

    def __init__(self, *parts):
        if isinstance(parts[0], DavPath):
            self.parts = parts[0].parts
        else:
            self.parts = parts

    def __call__(self, *parts):
        return self.__class__(*(self.parts + parts))

    def __repr__(self):
        return f'DavPath: {self.path}'

    @property
    def name(self):
        return self.parts[-1]

    @property
    def parent(self):
        return self.__class__(*self.parts[:-1])

    @property
    def check(self):
        return self.client.check(self.path)

    @property
    def version(self):
        return file_version(self)

    def increment_version(self):
        return increment_file_version(self)

    @property
    def info(self):
        return pipe(
            info(self.client, self), lambda info: merge(
                info, {
                    'modified':
                    (info['modified'] and parse_dt(info['modified'])),
                    'size': maybe_float(info['size']),
                }))

    _client = None
    @property
    def client(self):
        if self._client is None:
            print(self.__class__.client_function)
            self._client = self.__class__.client_function()
        return self._client

    @property
    def path_obj(self):
        return PurePosixPath(*self.parts)

    @property
    def path(self):
        return str(self.path_obj)

    @property
    def is_dir(self):
        return self.client.is_dir(self.path)

    def ls(self):
        return pipe(
            self.client.list(self.path),
            map(self),
            tuple,
        )

    def copy(self, to):
        return copy(self.client, self, to)

    def glob(self, glob_str):
        return glob(self.client, self, glob_str)

    def download(self, to: Path):
        return download(self.client, self, to)

    def load_workbook(self):
        return load_workbook(self.client, self)

    def upload_workbook(self, wb: 'Workbook'):
        return upload_workbook(self.client, wb, self)

    def load_binary(self):
        return load_binary(self.client, self)

    def load_text(self):
        return load_text(self.client, self)


version_re = re.compile(r'v(\d+)\.(\d+)')


def file_version(path: DavPath):
    if version_re.search(path.path):
        return maybe_pipe(
            path.path,
            version_re.findall,
            first,
            map(int),
            tuple,
        )


def increment_file_version(path: DavPath):
    if not version_re.search(path.path):
        log.error(f'DavPath {path} has no version. Adding v0.0')
        stem, ext = path.path_obj.stem, path.path_obj.suffix
        return path.__class__(*concatv(path.parent.parts, [f'{stem}-v0.0{ext}']))

    major, minor = path.version
    stem = version_re.sub(f'v{major}.{minor + 1}', path.path_obj.stem)
    ext = path.path_obj.suffix
    return path.__class__(*concatv(path.parent.parts, [f'{stem}{ext}']))


@curry
def copy(client: Client, from_path: DavPath, to_path: DavPath):
    return client.copy(from_path.path, to_path.path)


@curry
def download(client: Client, from_path: DavPath, to_path: Path):
    print(client)
    return client.download(from_path.path, str(to_path))


@curry
def upload(client: Client, from_path: Path, to_path: DavPath):
    return client.upload(to_path.path, str(from_path))


@curry
def load_workbook(client: Client, path: DavPath):
    import openpyxl
    with tempfile.TemporaryDirectory() as tempdir:
        wb_path = Path(tempdir, 'wb.xlsx')
        download(client, path, wb_path)
        wb = openpyxl.load_workbook(str(wb_path))
    return wb


@curry
def upload_workbook(client: Client, wb: 'Workbook', path: DavPath):
    with tempfile.TemporaryDirectory() as tempdir:
        wb_path = Path(tempdir, 'wb.xlsx')
        wb.save(str(wb_path))
        upload(client, wb_path, path)


@curry
def load_binary(client: Client, path: DavPath):
    with tempfile.TemporaryDirectory() as tempdir:
        temp_path = Path(tempdir, 'tempdata')
        download(client, path, temp_path)
        content = temp_path.read_bytes()
    return content


@curry
def load_text(client: Client, path: DavPath):
    with tempfile.TemporaryDirectory() as tempdir:
        temp_path = Path(tempdir, 'tempdata')
        download(client, path, temp_path)
        content = temp_path.read_text()
    return content


@curry
def glob(client: Client, from_path: DavPath, glob_str: str):
    glob_re = pipe(
        glob_str.replace('*', '.*').replace('?', '.?'),
        re.compile,
    )
    return pipe(
        client.list(from_path.path),
        filter(glob_re.search),
        map(from_path),
        tuple,
    )


@curry
def ls(client: Client, path: DavPath):
    return pipe(
        client.list(path.path),
        map(path),
        tuple,
    )


@curry
def info(client: Client, path: DavPath):
    return client.info(path.path)
