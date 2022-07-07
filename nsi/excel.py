from pathlib import Path

# Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.marker import DataPoint


# toolz
from . import logging
from .toolz import curry, ensure_paths, compose, escape_row

log = logging.new_log(__name__)

def load_workbook(path, *a, **kw) -> Workbook:
    '''Adds opened_from to Workbook object
    '''
    wb = openpyxl.load_workbook(path, *a, **kw)
    wb.opened_from = path
    return wb

@curry
def set_border(cell, sides={'top', 'left', 'right', 'bottom'}):
    kwargs = {s: Side(border_style='thin', color='FF000000')
              for s in sides}
    cell.border = Border(**kwargs)
    return cell

@curry
def set_alignment(cell, **kwargs):
    cell.alignment = Alignment(**kwargs)
    return cell

@curry
def set_fill(cell, **kwargs):
    cell.fill = PatternFill(**kwargs)
    return cell

set_gray = set_fill(patternType='solid', fgColor='d4d4d4')
set_green = set_fill(patternType='solid', fgColor='a9fca9')

@curry
def set_font(cell, **kwargs):
    cell.font = Font(**kwargs)
    return cell

@curry
def set_cell(sheet, ref, value):
    sheet[ref] = value
    return sheet[ref]

set_bold = set_font(b=True)

set_header = compose(set_border, set_bold, set_gray, set_cell)

@ensure_paths
def set_cells_all_sheets(workbook_path: Path, values: dict):
    wb = load_workbook(workbook_path)
    for sheet in wb:
        for ref, value in values.items():
            set_cell(sheet, ref, value)
    return wb
